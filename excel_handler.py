import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from config import IMPACT_LABELS, MAX_MATERIALS

COLOR_HEADER     = "1F4E79"
COLOR_REQUIRED   = "D9E1F2"
COLOR_OPTIONAL   = "EEF3F9"
COLOR_REF_HEADER = "2E7D32"
COLOR_API_HEADER = "4A235A"
COLOR_API_BG     = "F5EEF8"
COLOR_OUTPUT_OK  = "E8F5E9"
COLOR_OUTPUT_ERR = "FFCCCC"
COLOR_OUTPUT_FALL= "FFF9C4"


def generate_template(materials: list, products: list, countries: list) -> str:
    wb = Workbook()
    _create_ref_materials(wb, materials)
    _create_ref_products(wb, products)
    _create_ref_countries(wb, countries)
    _create_saisie_sheet(wb, materials, products, countries)
    _create_api_input_sheet(wb)
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.move_sheet("SAISIE", offset=-wb.sheetnames.index("SAISIE"))
    path = "ecobalyse_template.xlsx"
    wb.save(path)
    return path


def _ref_header(ws, headers):
    for col, h in enumerate(headers, start=1):
        c = ws.cell(1, col, h)
        c.font = Font(bold=True, color="FFFFFF", name="Arial")
        c.fill = PatternFill("solid", fgColor=COLOR_REF_HEADER)
        c.alignment = Alignment(horizontal="center")


def _create_ref_materials(wb, materials):
    ws = wb.create_sheet("REF_MATERIALS")
    _ref_header(ws, ["Nom (lisible)", "UUID (API)"])
    for i, m in enumerate(materials, start=2):
        ws.cell(i, 1, m["name"])
        ws.cell(i, 2, m["id"])
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 38


def _create_ref_products(wb, products):
    ws = wb.create_sheet("REF_PRODUCTS")
    _ref_header(ws, ["Nom (lisible)", "ID (API)"])
    for i, p in enumerate(products, start=2):
        ws.cell(i, 1, p["name"])
        ws.cell(i, 2, p["id"])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20


def _create_ref_countries(wb, countries):
    ws = wb.create_sheet("REF_COUNTRIES")
    _ref_header(ws, ["Nom (lisible)", "Code ISO (API)"])
    for i, c in enumerate(countries, start=2):
        ws.cell(i, 1, c["name"])
        ws.cell(i, 2, c["code"])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 15


def _get_saisie_columns():
    cols = [
        ("product_name",  "Nom du produit",  True),
        ("product_label", "Catégorie",        True),
        ("mass",          "Masse (kg)",        True),
    ]
    for i in range(1, MAX_MATERIALS + 1):
        req = (i == 1)
        cols += [
            (f"mat{i}_label",         f"Matière {i}",         req),
            (f"mat{i}_share",         f"Matière {i} — Part",  req),
            (f"mat{i}_country_label", f"Matière {i} — Pays",  False),
        ]
    cols += [
        ("countrySpinning_label", "Pays Filature",      False),
        ("countryFabric_label",   "Pays Tissage",        False),
        ("countryDyeing_label",   "Pays Ennoblissement", False),
        ("countryMaking_label",   "Pays Confection",     False),
    ]
    return cols


def _create_saisie_sheet(wb, materials, products, countries):
    ws = wb.create_sheet("SAISIE")
    columns = _get_saisie_columns()
    nb_mat     = len(materials) + 1
    nb_prod    = len(products)  + 1
    nb_country = len(countries) + 1

    for col_idx, (_, label, required) in enumerate(columns, start=1):
        c = ws.cell(1, col_idx, label)
        c.font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
        c.fill = PatternFill("solid", fgColor=COLOR_HEADER)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        fill = COLOR_REQUIRED if required else COLOR_OPTIONAL
        for row in range(2, 102):
            ws.cell(row, col_idx).fill = PatternFill("solid", fgColor=fill)
            ws.cell(row, col_idx).font = Font(name="Arial", size=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = 30 if "label" in _ or _ == "product_name" else 18

    ws.row_dimensions[1].height = 45
    ws.freeze_panes = "A2"

    col_map = {col[0]: i + 1 for i, col in enumerate(columns)}

    def add_dv(col_id, formula):
        if col_id not in col_map:
            return
        col = get_column_letter(col_map[col_id])
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}101")

    add_dv("product_label", f"REF_PRODUCTS!$A$2:$A${nb_prod}")
    for i in range(1, MAX_MATERIALS + 1):
        add_dv(f"mat{i}_label",         f"REF_MATERIALS!$A$2:$A${nb_mat}")
        add_dv(f"mat{i}_country_label", f"REF_COUNTRIES!$A$2:$A${nb_country}")
    for field in ["countrySpinning_label", "countryFabric_label",
                  "countryDyeing_label",   "countryMaking_label"]:
        add_dv(field, f"REF_COUNTRIES!$A$2:$A${nb_country}")


def _create_api_input_sheet(wb):
    ws = wb.create_sheet("API_INPUT")
    saisie_cols = _get_saisie_columns()
    saisie_map  = {col[0]: i + 1 for i, col in enumerate(saisie_cols)}

    def S(col_id, row):
        return f"SAISIE!${get_column_letter(saisie_map[col_id])}{row}"

    api_columns = ["product_name", "product", "mass"]
    for i in range(1, MAX_MATERIALS + 1):
        api_columns += [f"mat{i}_id", f"mat{i}_share", f"mat{i}_country"]
    api_columns += ["countrySpinning", "countryFabric", "countryDyeing", "countryMaking"]

    for col_idx, col_id in enumerate(api_columns, start=1):
        c = ws.cell(1, col_idx, col_id)
        c.font = Font(bold=True, color="FFFFFF", size=9, name="Arial")
        c.fill = PatternFill("solid", fgColor=COLOR_API_HEADER)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = 22
    ws.row_dimensions[1].height = 40

    for row in range(2, 102):
        formulas = {
            "product_name": f"={S('product_name', row)}",
            "product":      f'=IFERROR(VLOOKUP({S("product_label",row)},REF_PRODUCTS!$A:$B,2,FALSE),"")',
            "mass":         f"={S('mass', row)}",
            "countrySpinning": f'=IFERROR(VLOOKUP({S("countrySpinning_label",row)},REF_COUNTRIES!$A:$B,2,FALSE),"")',
            "countryFabric":   f'=IFERROR(VLOOKUP({S("countryFabric_label",row)},REF_COUNTRIES!$A:$B,2,FALSE),"")',
            "countryDyeing":   f'=IFERROR(VLOOKUP({S("countryDyeing_label",row)},REF_COUNTRIES!$A:$B,2,FALSE),"")',
            "countryMaking":   f'=IFERROR(VLOOKUP({S("countryMaking_label",row)},REF_COUNTRIES!$A:$B,2,FALSE),"")',
        }
        for i in range(1, MAX_MATERIALS + 1):
            formulas[f"mat{i}_id"]      = f'=IFERROR(VLOOKUP({S(f"mat{i}_label",row)},REF_MATERIALS!$A:$B,2,FALSE),"")'
            formulas[f"mat{i}_share"]   = f"={S(f'mat{i}_share', row)}"
            formulas[f"mat{i}_country"] = f'=IFERROR(VLOOKUP({S(f"mat{i}_country_label",row)},REF_COUNTRIES!$A:$B,2,FALSE),"")'

        for col_idx, col_id in enumerate(api_columns, start=1):
            if col_id in formulas:
                c = ws.cell(row, col_idx, formulas[col_id])
                c.fill = PatternFill("solid", fgColor=COLOR_API_BG)
                c.font = Font(name="Arial", size=9, color="555555")

    ws.freeze_panes = "A2"


def read_input(file) -> list:
    df = pd.read_excel(file, sheet_name="API_INPUT", dtype=str)
    df = df.dropna(how="all")
    df = df[df["product_name"].notna() & (df["product_name"].str.strip() != "")]
    return df.to_dict(orient="records")


def write_output(rows_input: list, results: list) -> str:
    output_rows = []
    for row, result in zip(rows_input, results):
        out = {"product_name": row.get("product_name", "")}
        if "error" in result and not result.get("impacts"):
            out["statut"]        = "ERREUR"
            out["erreur_detail"] = str(result["error"])
            out["fallback_note"] = result.get("fallback_note", "")
        else:
            out["statut"]        = "OK"
            out["erreur_detail"] = ""
            out["fallback_note"] = result.get("fallback_note", "")
            impacts = result.get("impacts", {})
            for code, label in IMPACT_LABELS.items():
                out[label] = impacts.get(code, "")
        output_rows.append(out)

    df   = pd.DataFrame(output_rows)
    path = "ecobalyse_results.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="RESULTATS")
        ws = writer.sheets["RESULTATS"]
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", name="Arial")
            cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
        for row in ws.iter_rows(min_row=2):
            fgColor = COLOR_OUTPUT_ERR if row[1].value == "ERREUR" else (COLOR_OUTPUT_FALL if row[2].value else COLOR_OUTPUT_OK)
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=fgColor)
                cell.font = Font(name="Arial", size=10)
    return path